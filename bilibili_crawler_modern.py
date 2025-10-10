from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import time

# 创建浏览器实例
browser = webdriver.Chrome()
WAIT = WebDriverWait(browser, 10)  # 最多等待10秒
browser.set_window_size(1400, 900)  # 设置浏览器窗口大小

# 创建Excel工作簿和工作表
wb = Workbook()
ws = wb.active  # 获取活动工作表
ws.title = '蔡徐坤篮球'  # 设置工作表名称

# 定义表头样式
header_font = Font(name='Arial', size=12, bold=True)
header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')

# 定义列名和列宽
columns = [
    ('视频标题', 40),
    ('视频地址', 30),
    ('视频时长', 15),
    ('观看次数', 15),
    ('弹幕数', 15),
    ('发布时间', 20)
]

# 写入表头并设置列宽
for col_idx, (header, width) in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    # 设置表头样式
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    # 设置列宽
    ws.column_dimensions[chr(64 + col_idx)].width = width

row = 2  # 数据从第二行开始写入

def save_to_excel(soup):
    global row
    video_cards = soup.find(class_='search-layout clearfix').find_all(class_='bili-video-card')

    for card in video_cards:
        try:
            # 提取数据
            title = card.find('h3').get('title')
            link = "https:" + card.find('a').get('href')
            duration = card.find(class_='bili-video-card__stats__duration').text
            
            stats_left = card.find(class_='bili-video-card__stats--left')
            view_count = stats_left.find_all('span')[0].text.strip()
            danmu_count = stats_left.find_all('span')[3].text.strip()
            
            publish_date = card.find(class_='bili-video-card__info--date').text.strip()

            # 打印调试信息
            print(f"正在爬取: {title}")
            print(f"链接: {link}")
            print(f"时长: {duration}")
            print(f"播放量: {view_count}")
            print(f"弹幕数: {danmu_count}")
            print(f"发布时间: {publish_date}")

            # 写入Excel
            ws.cell(row=row, column=1, value=title)
            ws.cell(row=row, column=2, value=link)
            ws.cell(row=row, column=3, value=duration)
            ws.cell(row=row, column=4, value=view_count)
            ws.cell(row=row, column=5, value=danmu_count)
            ws.cell(row=row, column=6, value=publish_date)

            # 设置单元格样式
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            row += 1

        except Exception as e:
            print(f"处理视频卡片时出错: {str(e)}")
            continue

def get_source():
    print("正在获取页面信息...")
    try:
        # 等待搜索结果容器和视频卡片加载完成
        WAIT.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.search-content > div.search-page-wrapper')))
        WAIT.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.bili-video-card')))
        
        # 确保页面完全加载
        time.sleep(1)
        
        html = browser.page_source
        soup = BeautifulSoup(html, "lxml")
        save_to_excel(soup)
        
    except TimeoutException as e:
        print(f"页面加载超时: {str(e)}")
        browser.refresh()
        time.sleep(2)
        get_source()
    except Exception as e:
        print(f"获取数据时出错: {str(e)}")
        browser.refresh()
        time.sleep(2)
        get_source()

def search():
    try:
        print("开始访问bilibili...")
        browser.get("https://www.bilibili.com/")

        # 查找搜索框并输入关键词
        search_input = WAIT.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".nav-search-input")))
        search_input.send_keys('蔡徐坤 篮球')

        # 点击搜索按钮
        submit = WAIT.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.nav-search-btn')))
        submit.click()

        print("跳转到搜索结果页面...")
        # 切换到新打开的窗口
        all_handles = browser.window_handles
        browser.switch_to.window(all_handles[1])
        
        # 获取第一页数据
        get_source()

        # 获取总页数
        total_pages = WAIT.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, '.vui_pagenation > div.vui_pagenation--btns > button:nth-last-child(2)')))
        total = int(total_pages.text)
        print(f"总页数: {total}")
        return total

    except TimeoutException:
        return search()

def next_page(page_num):
    try:
        print(f'开始获取第{page_num}页数据...')
        # 等待页面完全加载
        WAIT.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.vui_pagenation')))
        
        # 确保下一页按钮可点击
        next_btn = WAIT.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, '.vui_pagenation > div.vui_pagenation--btns > button:last-child')))
        
        # 使用JavaScript点击按钮
        browser.execute_script("arguments[0].click();", next_btn)
        
        # 等待新页面加载完成
        WAIT.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.search-content')))
        WAIT.until(EC.text_to_be_present_in_element(
            (By.CSS_SELECTOR, '.vui_pagenation > div.vui_pagenation--btns > button.vui_button--active'), 
            str(page_num)))
        
        time.sleep(1)  # 短暂等待确保页面稳定
        get_source()

    except TimeoutException as e:
        print(f"超时错误: {str(e)}")
        browser.refresh()
        time.sleep(2)
        return next_page(page_num)
    except Exception as e:
        print(f"发生错误: {str(e)}")
        browser.refresh()
        time.sleep(2)
        return next_page(page_num)

def main():
    try:
        total = search()
        print(f'搜索结果总页数为: {total}')

        # 遍历剩余页面
        for i in range(2, total + 1):
            next_page(i)
    
    except Exception as e:
        print(f"程序执行出错: {str(e)}")
    finally:
        try:
            print("正在保存Excel文件...")
            wb.save('蔡徐坤篮球.xlsx')
            print("Excel文件保存成功！")
        except Exception as e:
            print(f"保存Excel文件时出错: {str(e)}")
        finally:
            browser.quit()  # 关闭浏览器

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n程序被用户中断")
    except Exception as e:
        print(f"程序发生错误: {str(e)}")
    finally:
        print("程序结束")